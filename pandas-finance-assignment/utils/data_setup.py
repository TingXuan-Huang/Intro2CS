"""Data acquisition for "Four Formats, One Analysis".

All Part-0 logic lives here so the notebook cells stay 1-3 lines. Public functions:

- download_fred_csv()       -> data/DGS10.csv          (CSV, saved raw, "." markers kept)
- write_portfolio_json()    -> data/portfolio.json     (nested JSON, deterministic)
- download_fred_excel()     -> data/fred_macro.xlsx    (multi-sheet Excel, FRED export style)
- download_stocks_parquet() -> data/stock_market.parquet (long/tidy OHLCV via yfinance)
- inject_chaos(stocks)      -> deterministically messes up the stock table (Part 2)
- use_offline_data()        -> copies dated data_offline snapshots into data/ when downloads fail
"""
from __future__ import annotations

import io
import json
import shutil
import urllib.request
from pathlib import Path

import numpy as np
import pandas as pd

FRED_CSV_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}&cosd={start}"

CANONICAL_STOCK_COLUMNS = ["Date", "Ticker", "Open", "High", "Low", "Close", "Volume"]

DEFAULT_TICKERS = ("AAPL", "MSFT", "SPY")


def _http_get(url: str, timeout: int = 60) -> bytes:
    """Fetch a URL with a browser-ish User-Agent (FRED rejects some default agents)."""
    req = urllib.request.Request(
        url, headers={"User-Agent": "Mozilla/5.0 (pandas-finance-assignment)"}
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def download_fred_csv(
    series: str = "DGS10", start: str = "2021-01-01", out: str | Path = "data/DGS10.csv"
) -> Path:
    """Download a FRED series as CSV and save it *raw* (no cleaning, no parsing).

    FRED marks missing values (market holidays) with a literal "." — that quirk is
    the whole point of Task 1.1, so it is deliberately preserved. If FRED ever
    starts returning blank cells instead, we reinsert "." so the exercise survives.
    FRED has used both "observation_date" and "DATE" as the date header over the
    years; we normalize it to "DATE".
    """
    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)

    text = _http_get(FRED_CSV_URL.format(series=series, start=start)).decode("utf-8-sig")
    lines = [ln.strip() for ln in text.strip().splitlines() if ln.strip()]
    header, data_lines = lines[0], lines[1:]

    cols = [c.strip().strip('"') for c in header.split(",")]
    if len(cols) != 2 or cols[1] != series:
        raise RuntimeError(
            f"Unexpected FRED response header {cols!r} — expected a date column and {series!r}."
        )
    cols[0] = "DATE"  # normalize; FRED has shipped both spellings

    fixed_lines = []
    for ln in data_lines:
        date_part, _, value = ln.partition(",")
        if value.strip().strip('"') in ("", "nan", "NaN", "NA", "#N/A"):
            value = "."  # preserve the Task 1.1 exercise
        fixed_lines.append(f"{date_part},{value}")

    if len(fixed_lines) < 500:
        raise RuntimeError(
            f"FRED returned only {len(fixed_lines)} data rows for {series} — "
            "expected at least 500. Try again, or use use_offline_data()."
        )

    out.write_text(",".join(cols) + "\n" + "\n".join(fixed_lines) + "\n")
    n_missing = sum(ln.endswith(",.") for ln in fixed_lines)
    print(f"Saved {out} — {len(fixed_lines)} rows, {n_missing} missing markers ('.').")
    return out


def write_portfolio_json(out: str | Path = "data/portfolio.json") -> Path:
    """Write the (deterministic) nested portfolio file a broker API would return.

    Invariant the grader relies on: per holding, shares == sum of trade shares.
    """
    portfolio = {
        "owner": "Student Investor",
        "currency": "USD",
        "opened": "2021-01-04",
        "holdings": [
            {
                "ticker": "AAPL",
                "shares": 50,
                "trades": [
                    {"date": "2021-01-04", "shares": 30, "price": 129.41},
                    {"date": "2022-06-15", "shares": 20, "price": 135.87},
                ],
            },
            {
                "ticker": "MSFT",
                "shares": 25,
                "trades": [
                    {"date": "2021-03-10", "shares": 25, "price": 232.42},
                ],
            },
            {
                "ticker": "SPY",
                "shares": 40,
                "trades": [
                    {"date": "2021-01-04", "shares": 40, "price": 368.79},
                ],
            },
        ],
    }
    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(portfolio, indent=2) + "\n")
    print(f"Saved {out} — {len(portfolio['holdings'])} holdings, nested trades inside.")
    return out


def _extract_ticker_frame(raw: pd.DataFrame, ticker: str) -> pd.DataFrame:
    """Pull one ticker's OHLCV columns out of a yfinance download.

    yfinance's column layout varies by version and by group_by: the ticker can sit
    on either MultiIndex level. Handle all layouts defensively.
    """
    if isinstance(raw.columns, pd.MultiIndex):
        if ticker in set(raw.columns.get_level_values(0)):
            return raw[ticker].copy()
        return raw.xs(ticker, axis=1, level=-1).copy()
    return raw.copy()  # single-ticker download: flat columns


def download_stocks_parquet(
    tickers: tuple[str, ...] = DEFAULT_TICKERS,
    start: str = "2021-01-01",
    out: str | Path = "data/stock_market.parquet",
) -> Path:
    """Download daily OHLCV for `tickers` and save a long/tidy Parquet file.

    Canonical schema: columns exactly ["Date","Ticker","Open","High","Low","Close",
    "Volume"], one row per ticker-day, Date=datetime64, Ticker=string, prices=float,
    Volume=int64.
    """
    import yfinance as yf

    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)

    raw = yf.download(
        list(tickers), start=start, auto_adjust=True, progress=False, group_by="ticker"
    )
    if raw is None or len(raw) == 0:
        raise RuntimeError("yfinance returned no data — try again, or use use_offline_data().")

    frames = []
    for t in tickers:
        sub = _extract_ticker_frame(raw, t).reset_index()
        sub = sub.rename(columns={sub.columns[0]: "Date"})
        sub["Ticker"] = t
        frames.append(sub[CANONICAL_STOCK_COLUMNS])

    stocks = pd.concat(frames, ignore_index=True)
    stocks["Date"] = pd.to_datetime(stocks["Date"])
    if stocks["Date"].dt.tz is not None:
        stocks["Date"] = stocks["Date"].dt.tz_localize(None)
    stocks = stocks.dropna(subset=["Open", "High", "Low", "Close", "Volume"])

    # Keep only dates on which every ticker traded, so per-ticker row counts match.
    per_date = stocks.groupby("Date")["Ticker"].nunique()
    full_dates = per_date[per_date == len(tickers)].index
    stocks = stocks[stocks["Date"].isin(full_dates)]

    stocks["Ticker"] = stocks["Ticker"].astype(str)
    for col in ("Open", "High", "Low", "Close"):
        stocks[col] = stocks[col].astype(float)
    stocks["Volume"] = stocks["Volume"].round().astype("int64")
    stocks = stocks.sort_values(["Ticker", "Date"]).reset_index(drop=True)

    stocks.to_parquet(out, index=False, engine="pyarrow")
    print(f"Saved {out} — {len(stocks)} rows, {stocks['Ticker'].nunique()} tickers.")
    return out


def _write_fred_style_xlsx(df: pd.DataFrame, out: str | Path) -> Path:
    """Write `df` (observation_date + series columns) in FRED's Excel export style.

    Sheet "Monthly": rows 1-3 metadata strings, row 4 blank, row 5 the real header,
    data below. Sheet "Notes": series id -> plain-English description.
    """
    from openpyxl import Workbook

    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)

    series_cols = [c for c in df.columns if c != "observation_date"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly"
    ws.append(["FRED (Federal Reserve Economic Data) | fred.stlouisfed.org"])
    ws.append(["Units: CPIAUCSL: Index 1982-1984=100, Seasonally Adjusted; UNRATE: Percent"])
    ws.append(["Frequency: Monthly"])
    ws.append([])  # row 4 blank
    ws.append(["observation_date"] + series_cols)  # row 5: the real header
    for _, row in df.iterrows():
        date = pd.Timestamp(row["observation_date"]).to_pydatetime()
        values = [None if pd.isna(row[c]) else float(row[c]) for c in series_cols]
        ws.append([date] + values)
    for cell in ws["A"]:
        cell.number_format = "YYYY-MM-DD"

    notes = wb.create_sheet("Notes")
    notes.append(["series_id", "description"])
    notes.append(["CPIAUCSL", "Consumer Price Index for All Urban Consumers: All Items"])
    notes.append(["UNRATE", "Unemployment Rate (share of the labor force without a job)"])

    wb.save(out)
    return out


def download_fred_excel(out: str | Path = "data/fred_macro.xlsx") -> Path:
    """Download CPI + unemployment from FRED and save as a multi-sheet .xlsx.

    FRED's website offers Excel export via a button, but there is no stable
    scriptable Excel URL — so we download the CSV and recreate FRED's export
    style faithfully (metadata rows, then the header, plus a Notes sheet).
    """
    text = _http_get(
        FRED_CSV_URL.format(series="CPIAUCSL,UNRATE", start="2021-01-01")
    ).decode("utf-8-sig")
    df = pd.read_csv(io.StringIO(text), na_values=".")
    df = df.rename(columns={df.columns[0]: "observation_date"})
    df["observation_date"] = pd.to_datetime(df["observation_date"])
    if len(df) < 48:
        raise RuntimeError(
            f"FRED returned only {len(df)} monthly rows — expected at least 48. "
            "Try again, or use use_offline_data()."
        )
    out = _write_fred_style_xlsx(df, out)
    print(f"Saved {out} — sheets: Monthly ({len(df)} rows), Notes.")
    return out


def inject_chaos(stocks: pd.DataFrame, seed: int = 42) -> pd.DataFrame:
    """Deterministically inject the four classic data problems, then shuffle.

    Exact quantities (the grading and verification logic depend on them):
      * 60 duplicated rows (sampled with the seed),
      * 80 rows with a lower-cased Ticker,
      * 25 rows with Close = NaN,
      * exactly 1 fat-finger: one AAPL Close multiplied by 100, at least 250 rows
        into AAPL so rolling stats exist around it.

    Raises RuntimeError if the input already looks chaosed (mixed-case tickers or
    duplicate ticker-days) — running it twice would corrupt the exercise.
    """
    missing = [c for c in CANONICAL_STOCK_COLUMNS if c not in stocks.columns]
    if missing:
        raise ValueError(f"stocks is missing expected columns: {missing}")
    if not stocks["Ticker"].str.isupper().all() or stocks.duplicated(
        subset=["Ticker", "Date"]
    ).any():
        raise RuntimeError(
            "inject_chaos already ran on this data (or it is pre-chaosed offline data) "
            "— running it twice would double the mess. Skip this cell and go clean it!"
        )

    rng = np.random.default_rng(seed)
    df = stocks.sort_values(["Ticker", "Date"]).reset_index(drop=True)
    n = len(df)

    # 1) Pick the fat-finger row first: an AAPL row with >=250 rows on both sides.
    aapl_pos = df.index[df["Ticker"] == "AAPL"]
    candidates = aapl_pos[250 : len(aapl_pos) - 250]
    fat_pos = int(rng.choice(candidates))
    # Protect the fat finger and its neighbors from the other injections, so the
    # spike is always detectable by a neighbor-ratio test after deduplication.
    # Each ticker's first row is protected too: a NaN there would survive a
    # forward-fill and make the cleaning task unwinnable.
    window = range(max(fat_pos - 2, 0), min(fat_pos + 3, n))
    protected_keys = {(df.at[i, "Ticker"], df.at[i, "Date"]) for i in window}
    first_rows = df.groupby("Ticker").head(1).index
    protected_keys |= {(df.at[i, "Ticker"], df.at[i, "Date"]) for i in first_rows}

    def unprotected(index: pd.Index) -> np.ndarray:
        keys = list(zip(df.loc[index, "Ticker"], df.loc[index, "Date"]))
        return index[[k not in protected_keys for k in keys]].to_numpy()

    # 2) +60 duplicated rows (copies of non-protected originals).
    dup_rows = rng.choice(unprotected(df.index), size=60, replace=False)
    df = pd.concat([df, df.loc[dup_rows]], ignore_index=True)

    # 3) 25 rows with Close = NaN (never in the protected window).
    nan_rows = rng.choice(unprotected(df.index), size=25, replace=False)
    df.loc[nan_rows, "Close"] = np.nan

    # 4) 80 rows with a lower-cased Ticker.
    lower_rows = rng.choice(df.index.to_numpy(), size=80, replace=False)
    df.loc[lower_rows, "Ticker"] = df.loc[lower_rows, "Ticker"].str.lower()

    # 5) The fat finger itself (original positions 0..n-1 survived the concat).
    df.loc[fat_pos, "Close"] = df.loc[fat_pos, "Close"] * 100

    # 6) Shuffle so nothing is conveniently ordered.
    df = df.sample(frac=1, random_state=seed).reset_index(drop=True)
    print(
        f"Chaos injected: {len(df)} rows now contain 60 duplicates, 80 lowercase "
        "tickers, 25 missing closes and 1 fat-finger price. Good luck."
    )
    return df


def use_offline_data(data_dir: str | Path = "data", offline_dir: str | Path = "data_offline") -> None:
    """Copy dated FRED/Yahoo snapshot files into ``data/`` when downloads fail.

    ``data_offline/manifest.json`` records when a snapshot was captured and its
    redistribution note. The generated broker portfolio remains an instructional
    fixture because public trade-level portfolios are not privacy-safe course data.
    """
    data_dir, offline_dir = Path(data_dir), Path(offline_dir)
    data_dir.mkdir(parents=True, exist_ok=True)
    names = ["DGS10.csv", "portfolio.json", "fred_macro.xlsx", "stock_market.parquet"]
    for name in names:
        src = offline_dir / name
        if not src.exists():
            raise FileNotFoundError(
                f"{src} is missing — regenerate it with: python utils/make_offline_data.py"
            )
        shutil.copy(src, data_dir / name)
        print(f"Copied {src} -> {data_dir / name}")
    manifest_path = offline_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text())
        print(
            "Offline snapshot captured:",
            manifest.get("captured_at_utc", "unknown date"),
            "|",
            manifest.get("market_terms", "see assignment README"),
        )
    print(
        "Offline data in place. Note: the offline stock file is already chaosed — "
        "skip the inject_chaos cell in Part 2."
    )
